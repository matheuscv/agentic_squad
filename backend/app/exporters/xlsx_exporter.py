"""Exportador XLSX de Contatos (Fase D — TASK-06 / RF-07 a RF-10).

Gera um workbook `openpyxl` em memoria (`BytesIO`) com uma unica
planilha contendo os mesmos campos do CSV.

Decisao (R4 do PRD): o workbook e construido em memoria — apropriado
para o volume esperado da Fase D (RNF-02 cita 50k linhas). Para volumes
muito maiores no futuro, considerar `openpyxl.Workbook(write_only=True)`
ou geracao streaming via xlsxwriter; deixamos o caminho atual simples e
suficiente para a entrega.
"""

from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.exporters.csv_exporter import CONTATOS_CSV_HEADERS
from app.models.contato import Contato

# Mime type oficial OOXML para .xlsx — RNF-08 (arquivo deve abrir em
# Excel/LibreOffice/Google Sheets sem aviso de tipo desconhecido).
CONTATOS_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _formatar_datetime(valor):
    """Devolve datetime nativo para o openpyxl (mantem tipo de data na celula).

    Quando passamos um `datetime` nativo ao openpyxl, ele grava como
    celula de data formatada — preferivel a string ISO porque o usuario
    pode aplicar filtros/ordenacao de data no Excel sem reconverter.
    `None` vira celula vazia.
    """
    return valor


def gerar_xlsx_contatos(contatos: Iterable[Contato]) -> io.BytesIO:
    """Constroi um workbook XLSX e devolve um `BytesIO` posicionado em 0.

    Uso:
        buffer = gerar_xlsx_contatos(items)
        StreamingResponse(buffer, media_type=CONTATOS_XLSX_MIME)
    """
    workbook = Workbook()
    sheet = workbook.active
    # Nome da aba — limitado a 31 chars pelo formato OOXML.
    sheet.title = "Contatos"

    # Cabecalho (linha 1).
    sheet.append(list(CONTATOS_CSV_HEADERS))

    # Dados (linha 2 em diante). Convertemos id para int para que
    # planilhas reconhecam como numero; demais campos como string/datetime.
    for contato in contatos:
        sheet.append(
            [
                int(contato.id) if contato.id is not None else None,
                contato.nome or "",
                contato.email or "",
                contato.telefone or "",
                contato.empresa or "",
                _formatar_datetime(contato.criado_em),
                _formatar_datetime(contato.atualizado_em),
            ]
        )

    # Largura inicial razoavel das colunas — melhora a UX ao abrir a
    # planilha sem precisar auto-ajustar manualmente. Valores empiricos
    # (sem necessidade de medir caractere a caractere — economiza CPU).
    larguras_padrao = (8, 24, 32, 18, 24, 22, 22)
    for indice, largura in enumerate(larguras_padrao, start=1):
        sheet.column_dimensions[get_column_letter(indice)].width = largura

    # Congela cabecalho — celula A2 trava a linha 1 ao rolar.
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    # IMPORTANTE: posiciona o ponteiro no inicio para que o
    # StreamingResponse leia os bytes corretamente.
    buffer.seek(0)
    return buffer
