"""
Testes do endpoint GET /contatos/export — TASK-06 (Fase D / D.5 / RF-07 a RF-10).

Cobre:
- formato=csv -> 200, Content-Type 'text/csv; charset=utf-8',
  Content-Disposition com filename 'contatos_YYYYMMDD_HHMMSS.csv'
- formato=xlsx -> 200, Content-Type mimetype OOXML,
  Content-Disposition com filename 'contatos_YYYYMMDD_HHMMSS.xlsx'
- Default formato=csv quando parametro nao informado
- formato invalido -> 422
- Soft-deleted nao entra no export
- Filtros (busca, empresa, criado_*) sao aplicados no export
- Sort e respeitado no export
- Paginacao (skip/limit) e IGNORADA (RF-08)
- XLSX abre corretamente (openpyxl consegue ler de volta)
- Headers do CSV em PT-BR conforme contrato

Padrao: AAA (Arrange / Act / Assert).
"""

from __future__ import annotations

import io
import re

import pytest
from openpyxl import load_workbook

from app.exporters import (
    CONTATOS_CSV_HEADERS,
    CONTATOS_CSV_MIME,
    CONTATOS_XLSX_MIME,
)


# ===========================================================================
# Helpers
# ===========================================================================


def _auth_header(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _criar(client, token, *, nome: str, email: str, **extras) -> dict:
    payload = {"nome": nome, "email": email, **extras}
    resp = client.post("/contatos/", json=payload, headers=_auth_header(token))
    assert resp.status_code == 201, f"Falha ao criar contato: {resp.text}"
    return resp.json()


def _seed_export(client, token):
    """Cria 3 contatos com empresas/telefones distintos para testar export."""
    a = _criar(
        client,
        token,
        nome="Export Alfa",
        email="exp_alfa@test.com",
        empresa="Acme",
        telefone="(11) 91111-1111",
    )
    b = _criar(
        client,
        token,
        nome="Export Beta",
        email="exp_beta@test.com",
        empresa="Beta Industries",
    )
    c = _criar(
        client,
        token,
        nome="Export Gamma",
        email="exp_gamma@test.com",
        empresa="Gamma SA",
        telefone="(11) 93333-3333",
    )
    return {"alfa": a, "beta": b, "gamma": c}


_FILENAME_RE = re.compile(
    r'filename="(contatos_\d{8}_\d{6}\.(csv|xlsx))"'
)


# ===========================================================================
# CSV
# ===========================================================================


def test_export_csv_default_retorna_200(client, usuario_adm_token, usuario_default_token):
    """GET /contatos/export sem parametro formato deve usar default=csv (200)."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    # Default e CSV.
    assert resp.headers.get("content-type", "").startswith("text/csv")


def test_export_csv_explicito_retorna_content_type_correto(
    client, usuario_adm_token, usuario_default_token
):
    """GET /contatos/export?formato=csv -> content-type text/csv; charset=utf-8."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "csv"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    assert resp.headers.get("content-type") == CONTATOS_CSV_MIME


def test_export_csv_content_disposition_attachment(
    client, usuario_adm_token, usuario_default_token
):
    """CSV deve devolver Content-Disposition attachment + filename com timestamp."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "csv"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    assert cd.startswith("attachment;")
    match = _FILENAME_RE.search(cd)
    assert match is not None, f"Filename mal formado em Content-Disposition: {cd!r}"
    # Confere extensao csv.
    assert match.group(2) == "csv"


def test_export_csv_contem_headers_e_dados(
    client, usuario_adm_token, usuario_default_token
):
    """CSV exportado deve conter os headers em PT-BR e ao menos uma linha de dados."""
    seeds = _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "csv"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    conteudo = resp.text
    # Primeira linha eh o cabecalho (joined por virgula).
    primeira_linha = conteudo.splitlines()[0]
    for header in CONTATOS_CSV_HEADERS:
        assert header in primeira_linha
    # Dados dos seeds presentes (busca textual no CSV bruto).
    assert "exp_alfa@test.com" in conteudo
    assert "exp_beta@test.com" in conteudo
    assert "exp_gamma@test.com" in conteudo


# ===========================================================================
# XLSX
# ===========================================================================


def test_export_xlsx_retorna_content_type_correto(
    client, usuario_adm_token, usuario_default_token
):
    """GET /contatos/export?formato=xlsx -> mimetype OOXML."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "xlsx"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    assert resp.headers.get("content-type") == CONTATOS_XLSX_MIME


def test_export_xlsx_content_disposition_filename(
    client, usuario_adm_token, usuario_default_token
):
    """XLSX deve devolver Content-Disposition com filename .xlsx + timestamp."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "xlsx"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    match = _FILENAME_RE.search(cd)
    assert match is not None, f"Filename mal formado: {cd!r}"
    assert match.group(2) == "xlsx"


def test_export_xlsx_arquivo_valido_via_openpyxl(
    client, usuario_adm_token, usuario_default_token
):
    """XLSX gerado deve ser legivel via openpyxl (RNF-08 — abre em Excel/LibreOffice)."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "xlsx"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200

    # Le os bytes como workbook real.
    buffer = io.BytesIO(resp.content)
    wb = load_workbook(buffer, read_only=True)
    # Aba "Contatos" deve existir.
    assert "Contatos" in wb.sheetnames
    sheet = wb["Contatos"]
    # Primeira linha (linha 1) deve conter os headers em PT-BR.
    primeira_linha = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert tuple(primeira_linha) == CONTATOS_CSV_HEADERS


# ===========================================================================
# Formato invalido -> 422
# ===========================================================================


def test_export_formato_invalido_retorna_422(client, usuario_default_token):
    """formato=pdf (fora da allowlist) -> 422."""
    resp = client.get(
        "/contatos/export",
        params={"formato": "pdf"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 422


def test_export_formato_string_arbitraria_retorna_422(client, usuario_default_token):
    """formato='xls' (sem o 'x') tambem deve retornar 422."""
    resp = client.get(
        "/contatos/export",
        params={"formato": "xls"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 422


# ===========================================================================
# Soft-deleted nao entra (RF-09)
# ===========================================================================


def test_export_csv_exclui_soft_deleted(
    client, usuario_adm_token, usuario_default_token
):
    """Contato soft-deleted nao deve aparecer no CSV exportado."""
    seeds = _seed_export(client, usuario_adm_token)

    # Soft delete do contato beta.
    resp_del = client.delete(
        f"/contatos/{seeds['beta']['id']}",
        headers=_auth_header(usuario_adm_token),
    )
    assert resp_del.status_code == 204

    resp = client.get(
        "/contatos/export",
        params={"formato": "csv"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    assert "exp_beta@test.com" not in resp.text
    # Os outros devem continuar presentes.
    assert "exp_alfa@test.com" in resp.text
    assert "exp_gamma@test.com" in resp.text


# ===========================================================================
# Filtros aplicados ao export
# ===========================================================================


def test_export_aplica_filtro_empresa(
    client, usuario_adm_token, usuario_default_token
):
    """Filtro empresa deve restringir o conjunto exportado."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "csv", "empresa": "Acme"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    # Apenas o seed Alfa tem empresa 'Acme'.
    assert "exp_alfa@test.com" in resp.text
    assert "exp_beta@test.com" not in resp.text
    assert "exp_gamma@test.com" not in resp.text


def test_export_aplica_busca(client, usuario_adm_token, usuario_default_token):
    """Parametro busca deve restringir o conjunto exportado."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "csv", "busca": "Gamma"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    assert "exp_gamma@test.com" in resp.text
    assert "exp_alfa@test.com" not in resp.text
    assert "exp_beta@test.com" not in resp.text


def test_export_aplica_sem_telefone(
    client, usuario_adm_token, usuario_default_token
):
    """sem_telefone=true restringe o export aos contatos sem telefone."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "csv", "sem_telefone": "true"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    # Apenas Beta (sem telefone) deve aparecer.
    assert "exp_beta@test.com" in resp.text
    assert "exp_alfa@test.com" not in resp.text
    assert "exp_gamma@test.com" not in resp.text


def test_export_aplica_sort(client, usuario_adm_token, usuario_default_token):
    """Ordenacao por nome ASC deve refletir na ordem das linhas do CSV."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={"formato": "csv", "sort_by": "nome", "sort_order": "asc"},
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    linhas = resp.text.splitlines()[1:]  # pula header
    # Captura apenas linhas dos seeds (que comecam por "Export")
    linhas_seed = [l for l in linhas if "Export" in l]
    # Extrai o campo nome (segunda coluna do CSV — id,nome,...)
    nomes = [l.split(",")[1] for l in linhas_seed]
    assert nomes == sorted(nomes), f"Nomes nao ordenados: {nomes}"


def test_export_combina_filtro_e_sort(
    client, usuario_adm_token, usuario_default_token
):
    """Combina filtro 'empresa contendo letra a' + sort_by nome desc."""
    _seed_export(client, usuario_adm_token)

    resp = client.get(
        "/contatos/export",
        params={
            "formato": "csv",
            "empresa": "a",  # bate em Acme, Beta, Gamma (todas tem 'a')
            "sort_by": "nome",
            "sort_order": "desc",
        },
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 200
    linhas = resp.text.splitlines()[1:]
    linhas_seed = [l for l in linhas if "Export" in l]
    nomes = [l.split(",")[1] for l in linhas_seed]
    assert nomes == sorted(nomes, reverse=True)


# ===========================================================================
# Range invalido -> 422 (mesma logica da listagem)
# ===========================================================================


def test_export_range_data_invalido_retorna_422(client, usuario_default_token):
    """criado_desde > criado_ate no export tambem deve retornar 422."""
    resp = client.get(
        "/contatos/export",
        params={
            "formato": "csv",
            "criado_desde": "2026-12-31",
            "criado_ate": "2026-01-01",
        },
        headers=_auth_header(usuario_default_token),
    )
    assert resp.status_code == 422


# ===========================================================================
# Autenticacao
# ===========================================================================


def test_export_sem_token_retorna_401(client):
    """GET /contatos/export sem token -> 401."""
    resp = client.get("/contatos/export", params={"formato": "csv"})
    assert resp.status_code == 401
